# Version 1.0.1_DEV of ETAP Python library file
# Author lross2k
# Email luisross2000@outlook.com

import openpyxl   # For Excel files (.xlsx)
import xlwt, xlrd # For old Excel files (.xls)
from datetime import datetime
import time, re, os # For utils

# Class with methods and fields related to generating and manipulating ETAP Lumped files
class Lump():
    # Empty initialization
    def __init__(self):
        self._last_modified = time.localtime()
        self._last_cache_xlwt = time.localtime()
        self._last_cache_openpyxl = time.localtime()
        self._wb = None
        self._xlwt_wb = None
        self._xlrd_wb = None
        self.date_str = "%d/%m/%Y"
        self._timestep = None # Y-d-m HH:MM:SS
        self._rows = [] # row column syntax
        self._gen_header_row()
    
    # Calculate time step used in the current object
    def _calc_step(self):
        self._timestep = [self._rows[2][9] - self._rows[1][9],
                     self._rows[2][10] - self._rows[1][10],
                     self._rows[2][11] - self._rows[1][11]]
    
    # Generate the HEADER of Lump file
    def _gen_header_row(self):
        self._rows = [['P(MW)','Q(Mvar)','PF%','V(p.u.)','Angle', 'Humidity','Temp C',
                       'Wind(m/s)','Irradiance(W/m^2)','Hour','Min','Seconds','Date']]

    # Allow string conversion of the Lump structure for functions like str(), print()
    def __str__(self):
        longest = 0
        for row in self._rows:
            for col in row:
                longest = len(str(col)) if len(str(col)) > longest else longest
        string = ''
        for row in self._rows:
            for col in row:
                if col != '':
                    string = string + str(col) + ' '*(longest - len(str(col)))
                else:
                    string = string + ' '*longest
            string = string + '\n'
        return(string)
    
    # Set the current values for all items in the rows
    def set_values(self, P, Q, PF, Hour, Min, Sec, Date, V=None, Angle=None, 
                   Humidity=None, Temp=None, Wind=None, Irradiance=None, date_str=None):
        data = [P, Q, PF, V, Angle, Humidity, Temp, Wind, 
                Irradiance, Hour, Min, Sec, Date]
        tmp_row = []
        if date_str:
            self.date_str = date_str
        for i in range(0, len(P)):
            for j in range(0, 13):
                if data[j] != None:
                    if j == 12:
                        if type(data[j][i]) == str:
                            tmp_row.append(datetime.strptime(data[j][i], self.date_str) if data[j][i] != None else '')
                        else:
                            tmp_row.append(data[j][i] if data[j][i] != None else '')
                    else:
                        tmp_row.append(data[j][i] if data[j][i] != None else '')
                else:
                    tmp_row.append('')
            self._rows.append(tmp_row[::])
            tmp_row = []
        self._calc_step()
    
    # Save the current data in an Excel file, either xlsx or xls
    def save(self, filename, writer='openpyxl'):
        if writer == 'openpyxl':
            if self._wb != None and  self._last_cache_openpyxl > self._last_modified:
                self._wb.save(filename)
            else:
                time.sleep(.1)
                self._write_openpyxl_wb()
                try:
                    self.save(filename)
                except PermissionError:
                    print("Couldn't write to file, close it before running the program")
        elif writer == 'xlwt':
            if self._xlwt_wb != None and self._last_cache_xlwt > self._last_modified:
                self._xlwt_wb.save(filename)
            else:
                time.sleep(.1)
                self._write_xlwt_wb()
                try:
                    self.save(filename, writer='xlwt')
                except PermissionError:
                    print("Couldn't write to file, close it before running the program")

    # Generate data for an xls file using xlwt
    def _write_xlwt_wb(self):
        self._xlwt_wb = xlwt.Workbook()
        ws = self._xlwt_wb.add_sheet('Sheet1')
        date_format = xlwt.XFStyle()
        date_format.num_format_str = 'dd/mm/yyyy'
        for i in range(0, len(self._rows)):
            for j in range(0, 13):
                if j == 12:
                    ws.write(i,j,self._rows[i][j],date_format)
                else:
                    ws.write(i,j,self._rows[i][j])
        self._last_cache_xlwt = time.localtime()
                
    # Generate data for an xlsx file using openpyxl
    def _write_openpyxl_wb(self):
        self._wb = openpyxl.Workbook()
        ws = self._wb.active
        for i in range(1, len(self._rows)+1):
            for j in range(1, 14):
                try:
                    ws.cell(row=i,column=j).value = self._rows[i-1][j-1]
                except TypeError:
                    print(self._rows)
        self._last_cache_openpyxl = time.localtime()
    
    def change_time_step(self, time_step):
        if not self._timestep:
            print("No data has been added yet")
            return
        _rows_copy = self._rows[::]
        self._gen_header_row() 
        factor = (self._timestep[0]*60 + self._timestep[1]) / (time_step[0]*60 + time_step[1])
        tmp_row = []
        if type(factor) == float and factor < 1.0:
            step = int(1/factor)
            for i in range(1, len(_rows_copy), step):
                for j in range(0, 13):
                    tmp_row.append(_rows_copy[i][j])
                self._rows.append(tmp_row[::])
                tmp_row = []
        else:
            factor = int(factor)
            for i in range(1, len(_rows_copy)):
                for k in range(0,factor):
                    for j in range(0, 13):
                        if j == 10:
                            tmp_row.append(time_step[1]*k)
                        else:
                            tmp_row.append(_rows_copy[i][j])
                    self._rows.append(tmp_row[::])
                    tmp_row = []
        self._timestep = time_step
        self._last_modified = time.localtime()

# Load lump data from an Excel file using openpyxl for xlsx files and xlrd for xls files
def load_lump(filename, reader='openpyxl', date_str=None):
    _lump = Lump()
    if reader == 'openpyxl':
        _lump._wb = openpyxl.load_workbook(filename)
        _ws = _lump._wb.active
        P = []; Q = []; PF = []; V = []; Angle = []; Humidity = []; Temp = []
        Wind = []; Irradiance = []; Hour = []; Min = []; Sec = []; Date = []
        for row in range(2, _ws.max_row+1):
            P.append(         _ws.cell(row=row, column=1).value)
            Q.append(         _ws.cell(row=row, column=2).value)
            PF.append(        _ws.cell(row=row, column=3).value)
            V.append(         _ws.cell(row=row, column=4).value)
            Angle.append(     _ws.cell(row=row, column=5).value)
            Humidity.append(  _ws.cell(row=row, column=6).value)
            Temp.append(      _ws.cell(row=row, column=7).value)
            Wind.append(      _ws.cell(row=row, column=8).value)
            Irradiance.append(_ws.cell(row=row, column=9).value)
            Hour.append(      _ws.cell(row=row, column=10).value)
            Min.append(       _ws.cell(row=row, column=11).value)
            Sec.append(       _ws.cell(row=row, column=12).value)
            Date.append(      _ws.cell(row=row, column=13).value)
        _lump.set_values(P,Q,PF,Hour,Min,Sec,Date,V,Angle,Humidity,
                         Temp,Wind,Irradiance,date_str=date_str)
    elif reader == 'xlrd':
        _lump._xlrd_wb = xlrd.open_workbook(filename)
        _ws = _lump._xlrd_wb.sheet_by_index(0)
        P = []; Q = []; PF = []; V = []; Angle = []; Humidity = []; Temp = []
        Wind = []; Irradiance = []; Hour = []; Min = []; Sec = []; Date = []
        for row in range(1, _ws.nrows):
            P.append(         _ws.cell_value(rowx=row, colx=0))
            Q.append(         _ws.cell_value(rowx=row, colx=1))
            PF.append(        _ws.cell_value(rowx=row, colx=2))
            V.append(         _ws.cell_value(rowx=row, colx=3))
            Angle.append(     _ws.cell_value(rowx=row, colx=4))
            Humidity.append(  _ws.cell_value(rowx=row, colx=5))
            Temp.append(      _ws.cell_value(rowx=row, colx=6))
            Wind.append(      _ws.cell_value(rowx=row, colx=7))
            Irradiance.append(_ws.cell_value(rowx=row, colx=8))
            Hour.append(      _ws.cell_value(rowx=row, colx=9))
            Min.append(       _ws.cell_value(rowx=row, colx=10))
            Sec.append(       _ws.cell_value(rowx=row, colx=11))
            Date.append(      _ws.cell_value(rowx=row, colx=12))
        _lump.set_values(P,Q,PF,Hour,Min,Sec,Date,V,Angle,Humidity,
                         Temp,Wind,Irradiance,date_str=date_str)
    else:
        print('Unsupported reader, aborting')    
        return
    return(_lump)

''' Higher level function, allows to change the timestep of a file while also managing
automatically most of the file format and folder structure dependent parameters '''
def _timeshift_generic(filename, folder='.', hour=0, minutes=0,
                       date_str="%Y-%d-%m %H:%M:%S", alt_date_str="%d/%m/%Y"):
    if hour != 0 and minutes != 0:
        print('Unsuported timestep operation, aborting')
        return
    if re.search('.xls$', filename):
        reader = 'xlrd'
    elif re.search('.xlsx$', filename):
        reader = 'openpyxl'
    else:
        print('File:',filename,'should have .xls or .xlsx extension, aborting')
        return
    try:
        _lump_ = load_lump(filename, reader=reader, date_str=date_str)
    except ValueError:
        try:
            _lump_ = load_lump(filename, reader=reader, date_str=alt_date_str)
        except ValueError:
            print('File:',filename,'uses a different date format, specify it with date_str parameter, aborting')
            return
    if hour != 0:
        if _lump_._timestep[0] == hour: 
            print('File:',filename,'already had given hour time step, aborting')
            return
        else:
            _lump_.change_time_step([hour,0,0])
    elif minutes != 0:
        if _lump_._timestep[1] == minutes: 
            print('File:',filename,'already had given minutes time step, aborting')
            return
        else:
            _lump_.change_time_step([0,minutes,0])
    try:
        os.mkdir(folder)
    except FileExistsError:
        pass
    if reader == 'openpyxl':
        filename = filename.replace('xlsx','xls')
    _lump_.save(folder+'/'+filename, writer='xlwt')

def timeshift_to_60_minutes(filename, date_str="%Y-%d-%m %H:%M:%S"):
    _timeshift_generic(filename, folder='60_MIN', hour=1, date_str=date_str)

def timeshift_to_15_minutes(filename, date_str="%Y-%d-%m %H:%M:%S"):
    _timeshift_generic(filename, minutes=15, folder='15_MIN', date_str=date_str)

def timeshift_to_30_minutes(filename, date_str="%Y-%d-%m %H:%M:%S"):
    _timeshift_generic(filename, minutes=30, folder='30_MIN', date_str=date_str)

def _glob_timeshift_generic(func, date_str="%Y-%d-%m %H:%M:%S"):
    files = []
    for file in os.listdir():
        if re.search('[.]xl.+$', file):
            files.append(file)
    for file in files:
        func(file, date_str=date_str)

def glob_timeshift_to_60_minutes(date_str="%Y-%d-%m %H:%M:%S"):
    _glob_timeshift_generic(timeshift_to_60_minutes, date_str=date_str)

def glob_timeshift_to_30_minutes(date_str="%Y-%d-%m %H:%M:%S"):
    _glob_timeshift_generic(timeshift_to_30_minutes, date_str=date_str)

def glob_timeshift_to_15_minutes(date_str="%Y-%d-%m %H:%M:%S"):
    _glob_timeshift_generic(timeshift_to_15_minutes, date_str=date_str)
