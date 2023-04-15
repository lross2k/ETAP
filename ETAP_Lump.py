import openpyxl

class ETAP_Lump():
    # Empty initialization
    def __init__(self):
        self._wb = openpyxl.Workbook()
        self._ws = self._wb.active
        self._timestep = [0, 15, 0] # HH:MM:SS
        self._rows = [] # row column syntax
        self._gen_header_row()
    
    def calc_step(self, col):
        self._ws.cell(row=3, column=col).value - self._ws.cell(row=2, column=col).value
    
    def change_step(self):
        print('WIP')
    
    def _gen_header_row(self):
        self._rows = [['P(MW)','Q(Mvar)','PF%','V(p.u.)','Angle', 'Humidity','Temp C',
                       'Wind(m/s)','Irradiance(W/m^2)','Hour','Min','Seconds','Date']]

    def __str__(self):
        longest = 0
        for row in self._rows:
            for col in row:
                longest = len(str(col)) if len(str(col)) > longest else longest
        string = ''
        for row in self._rows:
            for col in row:
                if col != '':
                    string = string + str(col) + ' '*(longest - len(str(col)))
                else:
                    string = string + ' '*longest
            string = string + '\n'
        return(string)
    
    def set_values(self, P, Q, PF, V, Angle, Humidity, Temp,
                    Wind, Irradiance, Hour, Min, Sec, Date):
        data = [P, Q, PF, V, Angle, Humidity, Temp, Wind, 
                Irradiance, Hour, Min, Sec, Date]
        tmp_row = []
        for i in range(0, len(P)):
            for j in range(0, 13):
                if data[j] != None:
                    tmp_row.append(data[j][i])
                else:
                    tmp_row.append('')
            self._rows.append(tmp_row[::])
            tmp_row = []
